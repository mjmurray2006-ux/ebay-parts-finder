import os
import io
import re
import csv
import time
import uuid
import json
import base64
import sqlite3
import threading
from datetime import datetime, timedelta, timezone

import requests
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from apscheduler.schedulers.background import BackgroundScheduler
import sendgrid
from sendgrid.helpers.mail import (
    Mail, Attachment, FileContent, FileName, FileType, Disposition,
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# ─── eBay credentials ─────────────────────────────────────────────────────────
EBAY_CLIENT_ID     = os.environ.get("EBAY_CLIENT_ID", "")
EBAY_CLIENT_SECRET = os.environ.get("EBAY_CLIENT_SECRET", "")
EBAY_ENV           = os.environ.get("EBAY_ENV", "production")

# ─── Email / SendGrid ─────────────────────────────────────────────────────────
SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
ALERT_EMAIL_TO   = os.environ.get("ALERT_EMAIL_TO",   "marcus@apd.co.uk")
ALERT_EMAIL_FROM = os.environ.get("ALERT_EMAIL_FROM", "alerts@apd.co.uk")

# ─── Watchlist / trusted sellers persistence ──────────────────────────────────
WATCHLIST_FILE       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "watchlist.json")
TRUSTED_SELLERS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trusted_sellers.json")
MIN_FEEDBACK_SCORE   = 1000

# ─── Marshall price list persistence ──────────────────────────────────────────
# Marshall's monthly sheets run to hundreds of thousands of rows, so — unlike the
# watchlist/trusted-sellers JSON files above — this uses SQLite rather than a
# read-whole-file/write-whole-file JSON blob.
PRICE_DB_FILE        = os.path.join(os.path.dirname(os.path.abspath(__file__)), "marshall_prices.db")
QUOTE_TEMPLATE_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "quote_request_template.xlsx")
_price_db_lock       = threading.Lock()

# ─── In-memory state ──────────────────────────────────────────────────────────
_token_cache    = {"token": None, "expires_at": 0}
_watchlist_lock = threading.Lock()
_scan_status    = {
    "last_ran":     None,
    "items_checked": 0,
    "alerts_found": 0,
    "email_sent":   False,
    "scanning":     False,
    "scan_started_at": None,
}
SCAN_STALE_SECONDS = 600  # if a scan has been "in progress" longer than this, treat it as crashed
_last_scan_results = []  # most recent scan's per-item results, in-memory, for the weekly digest


# ═══════════════════════════════════════════════════════════════════════════════
# eBay auth
# ═══════════════════════════════════════════════════════════════════════════════

def get_token():
    now = time.time()
    if _token_cache["token"] and now < _token_cache["expires_at"] - 60:
        return _token_cache["token"]

    url = (
        "https://api.sandbox.ebay.com/identity/v1/oauth2/token"
        if EBAY_ENV == "sandbox"
        else "https://api.ebay.com/identity/v1/oauth2/token"
    )
    credentials = base64.b64encode(
        f"{EBAY_CLIENT_ID}:{EBAY_CLIENT_SECRET}".encode()
    ).decode()
    resp = requests.post(
        url,
        headers={
            "Authorization": f"Basic {credentials}",
            "Content-Type":  "application/x-www-form-urlencoded",
        },
        data={
            "grant_type": "client_credentials",
            "scope":      "https://api.ebay.com/oauth/api_scope",
        },
        timeout=10,
    )
    resp.raise_for_status()
    data = resp.json()
    _token_cache["token"]      = data["access_token"]
    _token_cache["expires_at"] = now + data.get("expires_in", 7200)
    return _token_cache["token"]


# ═══════════════════════════════════════════════════════════════════════════════
# Watchlist / trusted-sellers helpers
# ═══════════════════════════════════════════════════════════════════════════════

def load_watchlist():
    if not os.path.exists(WATCHLIST_FILE):
        return []
    with open(WATCHLIST_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_watchlist(items):
    with open(WATCHLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2)


def load_trusted_sellers():
    if not os.path.exists(TRUSTED_SELLERS_FILE):
        return []
    try:
        with open(TRUSTED_SELLERS_FILE, "r", encoding="utf-8") as f:
            return [str(s).strip().lower() for s in json.load(f) if str(s).strip()]
    except (OSError, ValueError):
        return []


def save_trusted_sellers(sellers):
    with open(TRUSTED_SELLERS_FILE, "w", encoding="utf-8") as f:
        json.dump(sellers, f, indent=2)


# ═══════════════════════════════════════════════════════════════════════════════
# Marshall price list — import, storage, lookup
#
# Marshall's monthly sheets list RRP only (not APD's cost — that still has to be
# emailed to Marshall separately per part). There is no markup/margin calculation
# here: the sheet's price is stored and shown as-is as the RRP.
#
# Column layout varies month to month and isn't assumed fixed — headers are
# matched defensively against a list of known aliases (see COLUMN_ALIASES)
# rather than by fixed position.
#
# Duplicate part numbers (across files, or re-imports of the same file later)
# are resolved by "most recent import wins": each import UPSERTs by part_number,
# so newer data always overwrites older data for that part. When one import
# batch contains multiple files with an overlapping part number, files are
# processed in the order they were uploaded and the later file wins.
# ═══════════════════════════════════════════════════════════════════════════════

COLUMN_ALIASES = {
    "part_number":        ["part number", "part no", "partnumber", "part num"],
    "description":        ["description", "desc"],
    "rrp":                ["retail price", "rrp", "price", "list price", "trade price"],
    "discount_code":      ["discount code"],
    "surcharge":          ["surcharge"],
    "superseded_part_no": ["superseding part no", "superseded part no", "supersession", "superseded by"],
    "unit_of_issue":      ["unit of issue", "unit"],
    "lead_time":          ["lead time", "leadtime"],
}


def get_price_db():
    conn = sqlite3.connect(PRICE_DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_price_db():
    conn = get_price_db()
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS parts (
                part_number         TEXT PRIMARY KEY,
                description         TEXT,
                rrp                 REAL,
                discount_code       TEXT,
                surcharge           REAL,
                superseded_part_no  TEXT,
                unit_of_issue       TEXT,
                lead_time           TEXT,
                brand               TEXT,
                source_file         TEXT,
                imported_at         TEXT
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_parts_description ON parts(description)")
        conn.commit()
    finally:
        conn.close()


def _norm_header(h):
    return re.sub(r"\s+", " ", str(h or "").strip().lower()).rstrip(".")


def _resolve_columns(headers):
    """Map canonical field names to column indices by matching header text against
    COLUMN_ALIASES, rather than assuming a fixed column order/name. If the same
    field's header appears more than once (Marshall's Jaguar sheets duplicate the
    'Superseding part no.' column), the first occurrence wins."""
    normalized = [_norm_header(h) for h in headers]
    resolved = {}
    for field, aliases in COLUMN_ALIASES.items():
        for idx, h in enumerate(normalized):
            if h in aliases:
                resolved[field] = idx
                break
    return resolved


def _safe_float(val):
    if val in (None, ""):
        return None
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return None


def _extract_row(row, cols):
    def get(field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        if val is None:
            return None
        return str(val).strip()

    part_number = get("part_number")
    if not part_number:
        return None

    return {
        "part_number":        part_number.upper(),
        "description":        get("description") or "",
        "rrp":                _safe_float(get("rrp")),
        "discount_code":      get("discount_code") or "",
        "surcharge":          _safe_float(get("surcharge")),
        "superseded_part_no": (get("superseded_part_no") or "").upper() or None,
        "unit_of_issue":      get("unit_of_issue") or "",
        "lead_time":          get("lead_time") or "",
    }


def _parse_csv_price_sheet(raw_bytes):
    text   = raw_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        headers = next(reader)
    except StopIteration:
        raise ValueError("File is empty")

    cols = _resolve_columns(headers)
    if "part_number" not in cols or "rrp" not in cols:
        raise ValueError(f"Could not find a Part Number / Retail Price column. Headers found: {headers}")

    for row in reader:
        if not row or not any(c.strip() for c in row if c):
            continue
        rec = _extract_row(row, cols)
        if rec:
            yield rec


def _parse_xlsx_price_sheet(raw_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        raise ValueError("File is empty")

    cols = _resolve_columns(headers)
    if "part_number" not in cols or "rrp" not in cols:
        raise ValueError(f"Could not find a Part Number / Retail Price column. Headers found: {list(headers)}")

    for row in rows_iter:
        if row is None or not any(row):
            continue
        rec = _extract_row(row, cols)
        if rec:
            yield rec


def parse_price_sheet(raw_bytes, filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".csv", ".txt"):
        return list(_parse_csv_price_sheet(raw_bytes))
    if ext in (".xlsx", ".xlsm"):
        return list(_parse_xlsx_price_sheet(raw_bytes))
    if ext == ".xls":
        raise ValueError("Legacy .xls format isn't supported — please save as .xlsx or .csv and re-upload.")
    raise ValueError(f"Unsupported file type '{ext}' — expected .csv or .xlsx")


def guess_brand(filename):
    stem = os.path.splitext(os.path.basename(filename))[0].strip().upper()
    if stem.startswith("JPF"):
        return "Jaguar"
    if stem.startswith("LPF"):
        return "Land Rover"
    return ""


init_price_db()


# ═══════════════════════════════════════════════════════════════════════════════
# Pricing logic
# ═══════════════════════════════════════════════════════════════════════════════

def get_confidence(listing_count):
    if listing_count >= 10:
        return "High"
    if listing_count >= 5:
        return "Medium"
    return "Low"


def calculate_action(your_price, competitor_lowest, threshold_percent):
    """Return (action, suggested_price, reason). competitor_lowest is None when out of stock."""
    if competitor_lowest is None:
        return "Raise", round(your_price, 2), "Trusted sellers out of stock — opportunity to raise price"

    if your_price > competitor_lowest * (1 + threshold_percent / 100):
        pct = (your_price - competitor_lowest) / your_price * 100
        return (
            "Lower",
            round(competitor_lowest * 1.02, 2),
            f"You are {pct:.1f}% above the trusted-seller lowest price",
        )

    if your_price < competitor_lowest * (1 - threshold_percent / 100):
        pct = (competitor_lowest - your_price) / competitor_lowest * 100
        return (
            "Raise",
            round(competitor_lowest * 0.95, 2),
            f"You are {pct:.1f}% below the trusted-seller lowest price — room to raise",
        )

    return "Hold", round(your_price, 2), f"Within {threshold_percent:.0f}% of the trusted-seller lowest price"


def calculate_price_trend(price_history):
    """rising only if each of the last 3 entries is strictly higher than the last,
    falling only if each is strictly lower, stable otherwise (incl. <3 entries)."""
    if len(price_history) < 3:
        return "stable"
    last3 = [h["price"] for h in price_history[-3:]]
    if last3[0] < last3[1] < last3[2]:
        return "rising"
    if last3[0] > last3[1] > last3[2]:
        return "falling"
    return "stable"


def price_n_days_ago(price_history, days=7):
    if not price_history:
        return None
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
    candidates = [h for h in price_history if h["date"] <= cutoff]
    return candidates[-1]["price"] if candidates else price_history[0]["price"]


def _ebay_base_url():
    return "https://api.sandbox.ebay.com" if EBAY_ENV == "sandbox" else "https://api.ebay.com"


def scan_watchlist_item(item):
    """
    Search eBay for this part exactly like /search (q = brand + part_number,
    relevance sort, limit 10), then filter to trusted sellers (from
    trusted_sellers.json) with 1000+ feedback and New condition only. Returns
    a dict describing what was found; raises only on a hard eBay/auth failure,
    which the caller records as a per-item scan error.
    """
    token = get_token()
    q = f"{item.get('brand', '')} {item.get('part_number', '')}".strip()
    resp = requests.get(
        f"{_ebay_base_url()}/buy/browse/v1/item_summary/search",
        headers={
            "Authorization":           f"Bearer {token}",
            "X-EBAY-C-MARKETPLACE-ID": item.get("market", "EBAY_GB"),
            "Content-Type":            "application/json",
        },
        # No sort param — relevance (eBay's default), matching /search.
        params={"q": q, "limit": 10},
        timeout=15,
    )
    resp.raise_for_status()

    trusted_sellers = set(load_trusted_sellers())
    raw_items       = resp.json().get("itemSummaries", [])

    listings = []
    for s in raw_items:
        try:
            price = float(s.get("price", {}).get("value", 0))
        except (ValueError, TypeError):
            continue
        if price <= 0:
            continue
        seller_obj = s.get("seller", {}) or {}
        seller     = (seller_obj.get("username") or "").lower()
        if seller not in trusted_sellers:
            continue
        try:
            feedback = int(seller_obj.get("feedbackScore") or 0)
        except (ValueError, TypeError):
            feedback = 0
        if feedback < MIN_FEEDBACK_SCORE:
            continue
        if (s.get("condition") or "").strip().lower() != "new":
            continue
        listings.append({"price": price, "seller": seller, "feedback": feedback})

    if not listings:
        return {
            "out_of_stock":          True,
            "trusted_sellers_found": 0,
            "listings_found":        len(raw_items),
            "competitor_lowest":     None,
        }

    return {
        "out_of_stock":          False,
        "trusted_sellers_found": len(listings),
        "listings_found":        len(raw_items),
        "competitor_lowest":     round(min(l["price"] for l in listings), 2),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Scan runner (called by routes and scheduler)
# ═══════════════════════════════════════════════════════════════════════════════

def run_scan(send_daily_email=True):
    global _scan_status, _last_scan_results
    started_at = _scan_status.get("scan_started_at")
    stale = started_at is not None and (time.time() - started_at) > SCAN_STALE_SECONDS
    if _scan_status["scanning"] and not stale:
        return {"error": "Scan already in progress"}, 409

    _scan_status["scanning"]        = True
    _scan_status["scan_started_at"] = time.time()
    lower_items  = []
    raise_items  = []
    clear_items  = []
    all_results  = []
    errors       = []
    alerts_count = 0

    try:
        with _watchlist_lock:
            items = load_watchlist()

        active = [it for it in items if it.get("active", True)]

        for idx, item in enumerate(active):
            was_out_of_stock = bool(item.get("competitor_out_of_stock"))
            try:
                scan = scan_watchlist_item(item)
            except Exception as e:
                errors.append({
                    "id":          item["id"],
                    "part_number": item.get("part_number", ""),
                    "error":       str(e),
                })
                if idx < len(active) - 1:
                    time.sleep(0.3)
                continue

            now_str    = datetime.now(timezone.utc).isoformat()
            your_price = item.get("your_price", 0)
            threshold  = item.get("alert_threshold_percent", 5)

            item["last_checked"]            = now_str
            item["competitor_out_of_stock"] = scan["out_of_stock"]

            if scan["out_of_stock"]:
                new_price = None
            else:
                new_price = scan["competitor_lowest"]
                item["last_price"] = new_price
                item.setdefault("price_history", []).append({"date": now_str[:10], "price": new_price})
                item["price_history"] = item["price_history"][-90:]
                item["price_trend"]   = calculate_price_trend(item["price_history"])

            action, suggested_price, reason = calculate_action(your_price, new_price, threshold)

            pct_diff = round((your_price - new_price) / new_price * 100, 1) if new_price else None
            price_changed = pct_diff is not None and abs(pct_diff) > threshold
            if price_changed:
                alerts_count += 1

            result = {
                "item":                  item,
                "competitor_lowest":     new_price,
                "listings_found":        scan["listings_found"],
                "trusted_sellers_found": scan["trusted_sellers_found"],
                "out_of_stock":          scan["out_of_stock"],
                "newly_out_of_stock":    scan["out_of_stock"] and not was_out_of_stock,
                "back_in_stock":         (not scan["out_of_stock"]) and was_out_of_stock,
                "action":                action,
                "suggested_price":       suggested_price,
                "reason":                reason,
                "confidence":            get_confidence(scan["trusted_sellers_found"]),
                "pct_diff":              pct_diff,
                "price_changed":         price_changed,
                "price_7d_ago":          price_n_days_ago(item.get("price_history", [])),
            }
            all_results.append(result)

            if action == "Lower":
                lower_items.append(result)
            elif action == "Raise":
                raise_items.append(result)
            else:
                clear_items.append(result)

            if idx < len(active) - 1:
                time.sleep(0.3)  # avoid hammering the eBay API

        with _watchlist_lock:
            save_watchlist(items)

        _last_scan_results = all_results

        summary = {
            "total_active":   len(active),
            "total_scanned":  len(all_results),
            "errors":         len(errors),
            "lower_needed":   len(lower_items),
            "raise_possible": len(raise_items),
            "hold":           len(clear_items),
            "alerts":         alerts_count,
        }

        email_sent, email_msg = (
            send_daily_alert_email(lower_items, raise_items, clear_items, summary, all_results)
            if send_daily_email else (False, "Daily email skipped")
        )

        _scan_status.update({
            "last_ran":      datetime.now(timezone.utc).isoformat(),
            "items_checked": len(all_results),
            "alerts_found":  alerts_count,
            "email_sent":    email_sent,
            "scanning":      False,
            "scan_started_at": None,
        })

        return {
            "summary": summary,
            "results": [
                {
                    "id":                r["item"]["id"],
                    "part_number":       r["item"].get("part_number", ""),
                    "brand":             r["item"].get("brand", ""),
                    "action":            r["action"],
                    "competitor_lowest": r["competitor_lowest"],
                    "suggested_price":   r["suggested_price"],
                    "confidence":        r["confidence"],
                    "price_changed":     r["price_changed"],
                    "out_of_stock":      r["out_of_stock"],
                }
                for r in all_results
            ],
            "errors":     errors,
            "email_sent": email_sent,
            "email_msg":  email_msg,
        }, 200

    except Exception as e:
        _scan_status["scanning"]        = False
        _scan_status["scan_started_at"] = None
        return {"error": str(e)}, 500


# ═══════════════════════════════════════════════════════════════════════════════
# Excel generation (MAM-compatible, two sheets)
# ═══════════════════════════════════════════════════════════════════════════════

TREND_ARROW = {"rising": "↑", "falling": "↓", "stable": "→"}

REPRICE_HEADERS = [
    "Part Number", "Brand", "Description", "Current Price (£)", "Suggested Price (£)",
    "Competitor Lowest (£)", "Difference (%)", "Trend", "Action", "Confidence", "Reason",
]
MARKET_HEADERS = REPRICE_HEADERS + [
    "Listings Found", "Trusted Sellers Found", "Last Checked", "Price 7 Days Ago", "Price Change This Week",
]

HDR_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
HDR_FONT = Font(bold=True)

ROW_FILLS = {
    "Lower": PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid"),
    "Raise": PatternFill(start_color="FFE8B3", end_color="FFE8B3", fill_type="solid"),
    "Hold":  PatternFill(start_color="D7FFD7", end_color="D7FFD7", fill_type="solid"),
}


def _write_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = Alignment(horizontal="center")


def _reprice_row(result):
    item = result["item"]
    return [
        item.get("part_number", ""),
        item.get("brand", ""),
        item.get("description", ""),
        item.get("your_price", 0),
        result.get("suggested_price"),
        result.get("competitor_lowest"),
        result.get("pct_diff"),
        TREND_ARROW.get(item.get("price_trend", "stable"), "→"),
        result.get("action", "Hold"),
        result.get("confidence", "Low"),
        result.get("reason", ""),
    ]


def generate_mam_excel(all_results):
    """Two-sheet MAM-compatible workbook: Reprice Actions (Lower/Raise only) and
    Full Market Report (every part scanned)."""
    wb = openpyxl.Workbook()

    # ── Sheet 1 — Reprice Actions ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Reprice Actions"
    _write_header(ws1, REPRICE_HEADERS)

    reprice_results = [r for r in all_results if r.get("action") in ("Lower", "Raise")]
    for row_idx, result in enumerate(reprice_results, 2):
        fill = ROW_FILLS.get(result.get("action"), ROW_FILLS["Hold"])
        for col_idx, val in enumerate(_reprice_row(result), 1):
            cell      = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            if col_idx in (4, 5, 6) and isinstance(val, (int, float)):
                cell.number_format = "£#,##0.00"

    for i, width in enumerate([16, 14, 30, 16, 18, 20, 14, 8, 10, 12, 50], 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # ── Sheet 2 — Full Market Report ─────────────────────────────────────────
    ws2 = wb.create_sheet("Full Market Report")
    _write_header(ws2, MARKET_HEADERS)

    for row_idx, result in enumerate(all_results, 2):
        item = result["item"]
        row_data = _reprice_row(result) + [
            result.get("listings_found", 0),
            result.get("trusted_sellers_found", 0),
            (item.get("last_checked") or "")[:10],
            result.get("price_7d_ago"),
            (
                round(result["competitor_lowest"] - result["price_7d_ago"], 2)
                if result.get("competitor_lowest") is not None and result.get("price_7d_ago") is not None
                else None
            ),
        ]
        fill = ROW_FILLS.get(result.get("action"), ROW_FILLS["Hold"])
        for col_idx, val in enumerate(row_data, 1):
            cell      = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            if col_idx in (4, 5, 6, 15, 16) and isinstance(val, (int, float)):
                cell.number_format = "£#,##0.00"

    for i, width in enumerate([16, 14, 30, 16, 18, 20, 14, 8, 10, 12, 50, 14, 16, 14, 16, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# Email — daily alert + weekly digest
# ═══════════════════════════════════════════════════════════════════════════════

def _email_table_rows(results, out_of_stock=False):
    rows = ""
    for r in results:
        item  = r["item"]
        trend = TREND_ARROW.get(item.get("price_trend", "stable"), "→")
        comp  = "Out of stock" if r.get("out_of_stock") else f"£{r.get('competitor_lowest', 0):.2f}"
        pct   = "—" if r.get("pct_diff") is None else f"{r['pct_diff']:.1f}%"
        rows += (
            f"<tr>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{item.get('part_number','')}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{item.get('brand','')}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>£{item.get('your_price',0):.2f}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{comp}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{pct}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{trend}</td>"
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee;font-weight:bold'>{r.get('action','Hold')}</td>"
            f"</tr>"
        )
    return rows


def _email_section(title, color, emoji, results):
    if not results:
        return ""
    return f"""
    <div style="margin-bottom:32px">
      <h2 style="color:{color};margin-bottom:8px">{emoji} {title}</h2>
      <table style="width:100%;border-collapse:collapse;font-size:13px">
        <thead>
          <tr style="background:{color};color:#fff">
            <th style="padding:8px 10px;text-align:left">Part Number</th>
            <th style="padding:8px 10px;text-align:left">Brand</th>
            <th style="padding:8px 10px;text-align:left">Your Price</th>
            <th style="padding:8px 10px;text-align:left">Competitor Lowest</th>
            <th style="padding:8px 10px;text-align:left">Diff %</th>
            <th style="padding:8px 10px;text-align:left">Trend</th>
            <th style="padding:8px 10px;text-align:left">Action</th>
          </tr>
        </thead>
        <tbody>{_email_table_rows(results)}</tbody>
      </table>
    </div>"""


def build_daily_email_html(lower_items, raise_items, clear_items, summary):
    today = datetime.now(timezone.utc).strftime("%d %B %Y")
    return f"""<!DOCTYPE html>
<html>
<body style="font-family:Arial,sans-serif;max-width:960px;margin:0 auto;padding:20px;color:#222">
  <h1 style="border-bottom:3px solid #1a5276;padding-bottom:10px;margin-bottom:24px">
    APD Price Alert
    <span style="display:block;font-size:14px;font-weight:normal;color:#555;margin-top:4px">{today}</span>
  </h1>

  {_email_section("Urgent — You're Being Undercut", "#c0392b", "🔴", lower_items)}
  {_email_section("Opportunity — Room To Raise", "#d68910", "🟡", raise_items)}

  <div style="background:#eafaf1;border-radius:8px;padding:20px;margin-top:24px">
    <h2 style="color:#1e8449;margin-top:0">🟢 All Clear</h2>
    <p style="margin:0;font-size:14px;color:#333">
      <strong>{len(clear_items)}</strong> part(s) checked with no action needed, out of
      <strong>{summary.get('total_scanned', 0)}</strong> scanned
      ({summary.get('errors', 0)} scan error(s)).
    </p>
  </div>

  <p style="margin-top:24px;font-size:12px;color:#999">
    MAM-compatible reprice spreadsheet attached (Reprice Actions + Full Market Report).
    Generated by eBay Parts Finder.
  </p>
</body>
</html>"""


def build_weekly_email_html(all_results):
    today          = datetime.now(timezone.utc)
    week_of        = today.strftime("%d %B %Y")
    changed        = [r for r in all_results if r.get("price_changed")]
    raise_ops      = [r for r in all_results if r.get("action") == "Raise" and not r.get("out_of_stock")]
    out_of_stock   = [r for r in all_results if r.get("out_of_stock")]
    back_in_stock  = [r for r in all_results if r.get("back_in_stock")]
    trend_counts   = {"rising": 0, "falling": 0, "stable": 0}
    for r in all_results:
        trend_counts[r["item"].get("price_trend", "stable")] = trend_counts.get(r["item"].get("price_trend", "stable"), 0) + 1

    return f"""<!DOCTYPE html>
<html>
<body style="font-family:Arial,sans-serif;max-width:960px;margin:0 auto;padding:20px;color:#222">
  <h1 style="border-bottom:3px solid #1a5276;padding-bottom:10px;margin-bottom:24px">
    APD Weekly Price Report
    <span style="display:block;font-size:14px;font-weight:normal;color:#555;margin-top:4px">Week of {week_of}</span>
  </h1>

  <div style="background:#eef3fb;border-radius:8px;padding:20px;margin-bottom:28px">
    <h2 style="color:#1a5276;margin-top:0">Overview</h2>
    <table style="font-size:14px;border-collapse:collapse">
      <tr><td style="padding:4px 20px 4px 0;color:#555">Parts monitored</td><td><strong>{len(all_results)}</strong></td></tr>
      <tr><td style="padding:4px 20px 4px 0;color:#555">Changed price this week</td><td><strong>{len(changed)}</strong></td></tr>
      <tr><td style="padding:4px 20px 4px 0;color:#555">Raise opportunities</td><td><strong style="color:#1e8449">{len(raise_ops)}</strong></td></tr>
      <tr><td style="padding:4px 20px 4px 0;color:#555">Trusted sellers currently out of stock</td><td><strong style="color:#d68910">{len(out_of_stock)}</strong></td></tr>
      <tr><td style="padding:4px 20px 4px 0;color:#555">New competitors appeared (back in stock)</td><td><strong>{len(back_in_stock)}</strong></td></tr>
      <tr><td style="padding:4px 20px 4px 0;color:#555">Price trend — rising / falling / stable</td>
          <td><strong>↑ {trend_counts.get('rising',0)} / ↓ {trend_counts.get('falling',0)} / → {trend_counts.get('stable',0)}</strong></td></tr>
    </table>
  </div>

  {_email_section("Changed Price This Week", "#1a5276", "📈", changed)}
  {_email_section("Raise Opportunities", "#1e8449", "🟢", raise_ops)}
  {_email_section("Trusted Sellers Out Of Stock — Opportunity To Raise", "#d68910", "🟠", out_of_stock)}
  {_email_section("New Competitors Appeared (Back In Stock)", "#8e44ad", "🆕", back_in_stock)}

  <p style="margin-top:24px;font-size:12px;color:#999">
    Full weekly MAM spreadsheet attached with suggested prices for every part.
    Generated by eBay Parts Finder.
  </p>
</body>
</html>"""


def _send_email(subject, html, all_results, filename_prefix):
    if not SENDGRID_API_KEY:
        return False, "SENDGRID_API_KEY not configured"

    excel_bytes   = generate_mam_excel(all_results)
    encoded_excel = base64.b64encode(excel_bytes).decode()
    today         = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    message = Mail(
        from_email=ALERT_EMAIL_FROM,
        to_emails=ALERT_EMAIL_TO,
        subject=subject,
        html_content=html,
    )
    message.attachment = Attachment(
        FileContent(encoded_excel),
        FileName(f"{filename_prefix}_{today}.xlsx"),
        FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        Disposition("attachment"),
    )

    try:
        sg       = sendgrid.SendGridAPIClient(api_key=SENDGRID_API_KEY)
        response = sg.send(message)
        return True, f"Sent (HTTP {response.status_code})"
    except Exception as e:
        return False, str(e)


def send_daily_alert_email(lower_items, raise_items, clear_items, summary, all_results):
    if not lower_items and not raise_items:
        return False, "No thresholds breached — daily email not sent"

    today   = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    needing = len(lower_items) + len(raise_items)
    subject = f"⚠️ APD Price Alert — {needing} part{'s' if needing != 1 else ''} need attention {today}"
    html    = build_daily_email_html(lower_items, raise_items, clear_items, summary)

    return _send_email(subject, html, all_results, "mam_reprice")


def send_weekly_digest_email(all_results):
    week_of = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    subject = f"📊 APD Weekly Price Report — week of {week_of}"
    html    = build_weekly_email_html(all_results)

    return _send_email(subject, html, all_results, "mam_weekly_report")


# ═══════════════════════════════════════════════════════════════════════════════
# Scheduler — daily scan 07:00 UTC, weekly digest Monday 08:00 UTC
# ═══════════════════════════════════════════════════════════════════════════════

def _scheduled_daily_scan():
    try:
        run_scan(send_daily_email=True)
    except Exception:
        pass


def _scheduled_weekly_digest():
    try:
        # The daily 07:00 UTC job already ran an hour earlier today, so reuse
        # its results rather than scanning again — falls back to a fresh scan
        # only if nothing recent is cached (e.g. first run after a restart).
        if not _last_scan_results:
            run_scan(send_daily_email=False)
        send_weekly_digest_email(_last_scan_results)
    except Exception:
        pass


SCHEDULER_LOCK_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".scheduler.lock")


def _acquire_scheduler_lock():
    """
    Guard against the daily/weekly jobs firing once per gunicorn worker if the
    worker count is ever increased beyond the current single-worker Procfile —
    only the process that wins this PID-file lock starts the scheduler. Stale
    locks (owning process no longer alive) are reclaimed automatically.
    """
    try:
        fd = os.open(SCHEDULER_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(os.getpid()).encode())
        os.close(fd)
        return True
    except FileExistsError:
        try:
            with open(SCHEDULER_LOCK_FILE, "r", encoding="utf-8") as f:
                old_pid = int(f.read().strip())
            os.kill(old_pid, 0)
            return False  # owning process is still alive
        except Exception:
            # Stale or unreadable lock — reclaim it.
            try:
                os.remove(SCHEDULER_LOCK_FILE)
            except OSError:
                pass
            try:
                fd = os.open(SCHEDULER_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
                os.write(fd, str(os.getpid()).encode())
                os.close(fd)
                return True
            except OSError:
                return False


_scheduler = BackgroundScheduler(daemon=True)

if _acquire_scheduler_lock():
    _scheduler.add_job(_scheduled_daily_scan, "cron", hour=7, minute=0, id="daily_scan", replace_existing=True)
    _scheduler.add_job(_scheduled_weekly_digest, "cron", day_of_week="mon", hour=8, minute=0, id="weekly_digest", replace_existing=True)
    _scheduler.start()


# ═══════════════════════════════════════════════════════════════════════════════
# Routes — health + eBay search (existing)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/search")
def search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "Missing search query"}), 400

    limit     = min(int(request.args.get("limit",  50)),  200)
    sort      = request.args.get("sort",      "relevance")
    min_price = request.args.get("min",       "")
    max_price = request.args.get("max",       "")
    condition = request.args.get("condition", "")
    market    = request.args.get("market",    "EBAY_GB")

    filters = []
    if min_price:
        filters.append(f"price:[{min_price}]")
    if max_price:
        filters.append(f"price:[..{max_price}]")
    if condition:
        filters.append(f"conditions:{{{condition}}}")

    params = {"q": q, "limit": limit}
    if sort != "relevance":
        params["sort"] = sort
    if filters:
        params["filter"] = ",".join(filters)

    try:
        token = get_token()
        resp  = requests.get(
            f"{_ebay_base_url()}/buy/browse/v1/item_summary/search",
            headers={
                "Authorization":           f"Bearer {token}",
                "X-EBAY-C-MARKETPLACE-ID": market,
                "Content-Type":            "application/json",
            },
            params=params,
            timeout=15,
        )
        resp.raise_for_status()
        return jsonify(resp.json())
    except requests.HTTPError as e:
        try:
            detail = e.response.json()
        except Exception:
            detail = {"raw": e.response.text}
        return jsonify({"error": str(e), "detail": detail}), e.response.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/seller")
def seller_search():
    username = request.args.get("username", "").strip()
    if not username:
        return jsonify({"error": "Missing username parameter"}), 400

    limit  = min(int(request.args.get("limit", 200)), 200)
    market = request.args.get("market", "EBAY_GB")

    # ── Attempt 1: Browse API seller filter ─────────────────────────────────
    try:
        token = get_token()
        resp  = requests.get(
            f"{_ebay_base_url()}/buy/browse/v1/item_summary/search",
            headers={
                "Authorization":           f"Bearer {token}",
                "X-EBAY-C-MARKETPLACE-ID": market,
                "Content-Type":            "application/json",
            },
            params={"q": username, "filter": f"sellers:{{{username}}}", "limit": limit},
            timeout=15,
        )
        if resp.ok:
            items = resp.json().get("itemSummaries", [])
            if items:
                return jsonify({"itemSummaries": items, "source": "browse_filter", "total": len(items)})
    except Exception:
        pass

    # ── Attempt 2: Finding API findItemsFromSeller ───────────────────────────
    global_id_map = {
        "EBAY_GB": "EBAY-GB", "EBAY_US": "EBAY-US",
        "EBAY_DE": "EBAY-DE", "EBAY_FR": "EBAY-FR", "EBAY_AU": "EBAY-AU",
    }
    finding_base = {
        "OPERATION-NAME":       "findItemsFromSeller",
        "SERVICE-VERSION":      "1.13.0",
        "SECURITY-APPNAME":     EBAY_CLIENT_ID,
        "RESPONSE-DATA-FORMAT": "JSON",
        "GLOBAL-ID":            global_id_map.get(market, "EBAY-GB"),
        "itemFilter(0).name":   "Seller",
        "itemFilter(0).value":  username,
        "sortOrder":            "CurrentPriceHighest",
    }
    try:
        all_items = []
        pages     = 2 if limit > 100 else 1
        for page in range(1, pages + 1):
            r = requests.get(
                "https://svcs.ebay.com/services/search/FindingService/v1",
                params={**finding_base, "paginationInput.pageSize": 100, "paginationInput.pageNumber": page},
                timeout=15,
            )
            if not r.ok:
                break
            result = r.json().get("findItemsFromSellerResponse", [{}])[0]
            if result.get("ack", [""])[0] != "Success":
                break
            raw = result.get("searchResult", [{}])[0].get("item", [])
            all_items.extend(_normalize_finding(i) for i in raw)
            if len(raw) < 100:
                break
        if all_items:
            return jsonify({"itemSummaries": all_items[:limit], "source": "finding_api", "total": len(all_items)})
    except Exception:
        pass

    # ── Attempt 3: Keyword fallback ──────────────────────────────────────────
    try:
        token = get_token()
        resp  = requests.get(
            f"{_ebay_base_url()}/buy/browse/v1/item_summary/search",
            headers={
                "Authorization":           f"Bearer {token}",
                "X-EBAY-C-MARKETPLACE-ID": market,
                "Content-Type":            "application/json",
            },
            params={"q": username, "limit": limit},
            timeout=15,
        )
        resp.raise_for_status()
        all_items = resp.json().get("itemSummaries", [])
        filtered  = [
            i for i in all_items
            if (i.get("seller", {}).get("username") or "").lower() == username.lower()
        ]
        return jsonify({"itemSummaries": filtered, "source": "keyword_fallback", "total": len(filtered)})
    except requests.HTTPError as e:
        try:
            detail = e.response.json()
        except Exception:
            detail = {"raw": e.response.text}
        return jsonify({"error": str(e), "detail": detail}), e.response.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _normalize_finding(item):
    price_obj  = item.get("sellingStatus", [{}])[0].get("currentPrice", [{}])[0]
    price_val  = price_obj.get("__value__", "0")
    price_cur  = price_obj.get("@currencyId", "GBP")
    cond       = item.get("condition",  [{}])[0].get("conditionDisplayName", [""])[0]
    seller_obj = item.get("seller",     [{}])[0]
    s_name     = seller_obj.get("sellerUserName", [""])[0]
    s_fb_raw   = seller_obj.get("feedbackScore",  ["0"])[0]
    s_fb       = int(s_fb_raw) if str(s_fb_raw).isdigit() else 0
    ship_obj   = item.get("shippingInfo", [{}])[0]
    ship_cost  = ship_obj.get("shippingServiceCost", [{}])[0].get("__value__", None)
    ship_cur   = ship_obj.get("shippingServiceCost", [{}])[0].get("@currencyId", price_cur)
    title      = item.get("title",       [""])[0]
    url        = item.get("viewItemURL", ["#"])[0]
    item_id    = item.get("itemId",      [""])[0]
    gallery    = item.get("galleryURL",  [""])[0]

    normalized = {
        "itemId":     item_id,
        "title":      title,
        "price":      {"value": price_val, "currency": price_cur},
        "condition":  cond,
        "seller":     {"username": s_name, "feedbackScore": s_fb},
        "itemWebUrl": url,
    }
    if ship_cost is not None:
        normalized["shippingOptions"] = [{"shippingCost": {"value": ship_cost, "currency": ship_cur}}]
    if gallery:
        normalized["image"] = {"imageUrl": gallery}
    return normalized


# ═══════════════════════════════════════════════════════════════════════════════
# Routes — watchlist CRUD
# (scan routes registered first so Flask prefers them over /<id> matches)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/watchlist/scan/status", methods=["GET"])
def scan_status():
    with _watchlist_lock:
        total_parts = len(load_watchlist())

    daily_job  = _scheduler.get_job("daily_scan")
    next_scan  = daily_job.next_run_time.isoformat() if daily_job and daily_job.next_run_time else None

    status = dict(_scan_status)
    status["total_parts"] = total_parts
    status["next_scan"]   = next_scan
    status["total_alerts"] = _scan_status.get("alerts_found", 0)
    return jsonify(status)


@app.route("/watchlist/scan", methods=["POST"])
def trigger_scan():
    started_at = _scan_status.get("scan_started_at")
    stale = started_at is not None and (time.time() - started_at) > SCAN_STALE_SECONDS
    if _scan_status["scanning"] and not stale:
        return jsonify({"error": "Scan already in progress"}), 409
    threading.Thread(target=run_scan, daemon=True).start()
    return jsonify({"started": True}), 202


@app.route("/watchlist", methods=["GET"])
def get_watchlist():
    with _watchlist_lock:
        items = load_watchlist()
    return jsonify(items)


def _build_watchlist_item(data):
    """Build a new watchlist item dict from request data. Raises KeyError/ValueError on bad input."""
    return {
        "id":                      str(uuid.uuid4()),
        "part_number":             str(data["part_number"]),
        "brand":                   str(data.get("brand", "")),
        "description":             str(data.get("description", "")),
        "your_price":              float(data["your_price"]),
        "alert_threshold_percent": float(data.get("alert_threshold_percent", 5)),
        "market":                  str(data.get("market", "EBAY_GB")),
        "active":                  bool(data.get("active", True)),
        "added_date":              datetime.now(timezone.utc).isoformat(),
        "last_checked":            None,
        "last_price":              None,
        "price_history":           [],
        "price_trend":             "stable",
        "competitor_out_of_stock": False,
    }


@app.route("/watchlist", methods=["POST"])
def add_watchlist_item():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    missing = [f for f in ("part_number", "your_price") if f not in data]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    try:
        item = _build_watchlist_item(data)
    except (ValueError, TypeError) as e:
        return jsonify({"error": f"Invalid field value: {e}"}), 400

    with _watchlist_lock:
        items = load_watchlist()
        items.append(item)
        save_watchlist(items)

    return jsonify(item), 201


@app.route("/watchlist/import", methods=["POST"])
def import_watchlist():
    data = request.get_json(silent=True)
    if not isinstance(data, list):
        return jsonify({"error": "JSON body must be an array of parts"}), 400

    imported = []
    errors   = []
    for i, entry in enumerate(data):
        if not isinstance(entry, dict) or "part_number" not in entry or "your_price" not in entry:
            errors.append({"index": i, "error": "Missing part_number or your_price"})
            continue
        try:
            imported.append(_build_watchlist_item(entry))
        except (ValueError, TypeError) as e:
            errors.append({"index": i, "error": str(e)})

    with _watchlist_lock:
        items = load_watchlist()
        items.extend(imported)
        save_watchlist(items)

    return jsonify({"imported": len(imported), "errors": errors, "items": imported}), 201


@app.route("/watchlist/<item_id>", methods=["PUT"])
def update_watchlist_item(item_id):
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"error": "JSON body required"}), 400

    numeric_fields = {"your_price", "alert_threshold_percent"}
    string_fields  = {"part_number", "brand", "description", "market"}
    allowed_fields = numeric_fields | string_fields | {"active"}

    with _watchlist_lock:
        items = load_watchlist()
        for i, item in enumerate(items):
            if item["id"] == item_id:
                for field in allowed_fields:
                    if field in data:
                        val = data[field]
                        if field in numeric_fields:
                            val = float(val)
                        elif field in string_fields:
                            val = str(val)
                        elif field == "active":
                            val = bool(val)
                        items[i][field] = val
                save_watchlist(items)
                return jsonify(items[i])

    return jsonify({"error": "Item not found"}), 404


@app.route("/watchlist/<item_id>", methods=["DELETE"])
def delete_watchlist_item(item_id):
    with _watchlist_lock:
        items    = load_watchlist()
        filtered = [it for it in items if it["id"] != item_id]
        if len(filtered) == len(items):
            return jsonify({"error": "Item not found"}), 404
        save_watchlist(filtered)
    return jsonify({"deleted": item_id})


@app.route("/watchlist/trusted-sellers", methods=["GET"])
def get_trusted_sellers_route():
    return jsonify(load_trusted_sellers())


@app.route("/watchlist/trusted-sellers", methods=["POST"])
def set_trusted_sellers_route():
    data = request.get_json(silent=True)
    if not isinstance(data, list):
        return jsonify({"error": "JSON body must be an array of seller usernames"}), 400

    sellers = sorted({str(s).strip().lower() for s in data if str(s).strip()})
    save_trusted_sellers(sellers)
    return jsonify({"trusted_sellers": sellers, "count": len(sellers)})


@app.route("/watchlist/export", methods=["GET"])
def export_watchlist_excel():
    with _watchlist_lock:
        items = load_watchlist()

    if _last_scan_results:
        results = _last_scan_results
    else:
        # No scan has run yet this process — export a "not yet scanned" sheet
        # rather than failing, so the button always produces something.
        results = [{
            "item":                  it,
            "competitor_lowest":     it.get("last_price"),
            "listings_found":        0,
            "trusted_sellers_found": 0,
            "out_of_stock":          bool(it.get("competitor_out_of_stock")),
            "action":                "Hold",
            "suggested_price":       it.get("your_price", 0),
            "reason":                "Not yet scanned",
            "confidence":            "Low",
            "pct_diff":              None,
            "price_changed":         False,
            "price_7d_ago":          price_n_days_ago(it.get("price_history", [])),
        } for it in items]

    excel_bytes = generate_mam_excel(results)
    today       = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    return send_file(
        io.BytesIO(excel_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"mam_reprice_{today}.xlsx",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Routes — Marshall price list
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/parts/import", methods=["POST"])
def import_price_sheets():
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "No files uploaded — attach one or more Marshall price sheets"}), 400

    brand_overrides = request.form.to_dict()  # optional: brand_<original filename> -> "Jaguar" etc.

    file_results = []
    with _price_db_lock:
        conn = get_price_db()
        try:
            for f in files:
                filename = f.filename or "upload"
                raw = f.read()
                try:
                    rows = parse_price_sheet(raw, filename)
                except ValueError as e:
                    file_results.append({"file": filename, "error": str(e), "rows_imported": 0})
                    continue

                brand = brand_overrides.get(f"brand_{filename}") or guess_brand(filename)
                now   = datetime.now(timezone.utc).isoformat()
                records = [
                    (
                        r["part_number"], r["description"], r["rrp"], r["discount_code"],
                        r["surcharge"], r["superseded_part_no"], r["unit_of_issue"], r["lead_time"],
                        brand, filename, now,
                    )
                    for r in rows
                ]

                if not records:
                    file_results.append({
                        "file": filename, "brand": brand, "rows_imported": 0,
                        "error": "No usable rows found in this file",
                    })
                    continue

                conn.executemany("""
                    INSERT INTO parts (part_number, description, rrp, discount_code, surcharge,
                                        superseded_part_no, unit_of_issue, lead_time, brand, source_file, imported_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(part_number) DO UPDATE SET
                        description=excluded.description,
                        rrp=excluded.rrp,
                        discount_code=excluded.discount_code,
                        surcharge=excluded.surcharge,
                        superseded_part_no=excluded.superseded_part_no,
                        unit_of_issue=excluded.unit_of_issue,
                        lead_time=excluded.lead_time,
                        brand=excluded.brand,
                        source_file=excluded.source_file,
                        imported_at=excluded.imported_at
                """, records)
                conn.commit()
                file_results.append({"file": filename, "brand": brand, "rows_imported": len(records)})

            total_parts = conn.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        finally:
            conn.close()

    return jsonify({"files": file_results, "total_parts": total_parts}), 200


@app.route("/parts/search", methods=["GET"])
def search_parts():
    q     = (request.args.get("q") or "").strip()
    limit = max(1, min(int(request.args.get("limit", 50) or 50), 200))
    if not q:
        return jsonify({"results": [], "total": 0})

    conn = get_price_db()
    try:
        q_upper = q.upper()
        exact = conn.execute(
            "SELECT * FROM parts WHERE part_number = ?", (q_upper,)
        ).fetchall()
        rest = conn.execute(
            """
            SELECT * FROM parts
            WHERE part_number != ? AND (part_number LIKE ? OR description LIKE ?)
            ORDER BY part_number
            LIMIT ?
            """,
            (q_upper, f"{q_upper}%", f"%{q}%", limit),
        ).fetchall()
        rows = (list(exact) + list(rest))[:limit]
        return jsonify({"results": [dict(r) for r in rows], "total": len(rows)})
    finally:
        conn.close()


@app.route("/parts/stats", methods=["GET"])
def parts_stats():
    conn = get_price_db()
    try:
        total       = conn.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        last_import = conn.execute("SELECT MAX(imported_at) FROM parts").fetchone()[0]
        by_brand    = conn.execute(
            "SELECT COALESCE(NULLIF(brand, ''), 'Unknown') AS brand, COUNT(*) AS n, MAX(imported_at) AS last "
            "FROM parts GROUP BY brand ORDER BY n DESC"
        ).fetchall()
        return jsonify({
            "total_parts": total,
            "last_import": last_import,
            "by_brand": [dict(r) for r in by_brand],
        })
    finally:
        conn.close()


@app.route("/parts/quote-request", methods=["POST"])
def generate_quote_request():
    data = request.get_json(silent=True)
    lines = data.get("lines") if isinstance(data, dict) else None
    if not isinstance(lines, list) or not lines:
        return jsonify({"error": "JSON body must include a non-empty 'lines' array"}), 400

    if not os.path.exists(QUOTE_TEMPLATE_FILE):
        return jsonify({"error": "Quote request template not found on the server"}), 500

    wb = openpyxl.load_workbook(QUOTE_TEMPLATE_FILE)
    ws = wb["Quote Request"] if "Quote Request" in wb.sheetnames else wb.active

    # Data starts at row 5 in the template; rows 1-4 are the title/instructions/
    # header and there's a footer note a little further down. Insert extra rows
    # if there are more lines than the template has room for, so a big quote
    # never overwrites the footer instead of just running past the pre-formatted area.
    TEMPLATE_DATA_ROWS = 47  # rows 5–51 in the stock template
    start_row = 5
    if len(lines) > TEMPLATE_DATA_ROWS:
        ws.insert_rows(start_row + TEMPLATE_DATA_ROWS, amount=len(lines) - TEMPLATE_DATA_ROWS)

    for i, line in enumerate(lines):
        if not isinstance(line, dict):
            continue
        row = start_row + i
        ws.cell(row=row, column=1, value=line.get("part_number") or "")
        ws.cell(row=row, column=2, value=line.get("description") or "")
        ws.cell(row=row, column=3, value=line.get("superseded_part_no") or "")
        ws.cell(row=row, column=4, value=line.get("qty") or 1)
        ws.cell(row=row, column=5, value=line.get("cost_price"))
        ws.cell(row=row, column=6, value=line.get("rrp"))
        ws.cell(row=row, column=7, value=line.get("lead_time") or "")
        ws.cell(row=row, column=8, value=line.get("notes") or "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"APD_Marshall_Quote_Request_{today}.xlsx",
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Routes — Taxonomy API (category search + item specifics)
# ═══════════════════════════════════════════════════════════════════════════════

_category_tree_cache = {}

REQUIREMENT_ORDER = {"REQUIRED": 0, "RECOMMENDED": 1, "OPTIONAL": 2}


def get_category_tree_id(marketplace):
    if marketplace in _category_tree_cache:
        return _category_tree_cache[marketplace]

    token = get_token()
    resp  = requests.get(
        f"{_ebay_base_url()}/commerce/taxonomy/v1/get_default_category_tree_id",
        headers={"Authorization": f"Bearer {token}"},
        params={"marketplace_id": marketplace},
        timeout=10,
    )
    resp.raise_for_status()
    tree_id = resp.json()["categoryTreeId"]
    _category_tree_cache[marketplace] = tree_id
    return tree_id


def _build_breadcrumb(category_name, ancestors):
    ordered = sorted(ancestors, key=lambda a: a.get("categoryTreeNodeLevel", 0))
    return " > ".join([a["categoryName"] for a in ordered] + [category_name])


@app.route("/api/taxonomy/search-categories", methods=["GET"])
def search_categories():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"error": "Missing search query"}), 400

    marketplace = request.args.get("marketplace", "EBAY_GB")

    try:
        tree_id = get_category_tree_id(marketplace)
        token   = get_token()
        resp    = requests.get(
            f"{_ebay_base_url()}/commerce/taxonomy/v1/category_tree/{tree_id}/get_category_suggestions",
            headers={"Authorization": f"Bearer {token}"},
            params={"q": q},
            timeout=15,
        )
        resp.raise_for_status()

        suggestions = []
        for s in resp.json().get("categorySuggestions", []):
            category = s.get("category", {})
            suggestions.append({
                "categoryId":   category.get("categoryId"),
                "categoryName": category.get("categoryName"),
                "breadcrumb":   _build_breadcrumb(
                    category.get("categoryName", ""),
                    s.get("categoryTreeNodeAncestors", []),
                ),
            })

        return jsonify(suggestions)
    except requests.HTTPError as e:
        try:
            detail = e.response.json()
        except Exception:
            detail = {"raw": e.response.text}
        return jsonify({"error": str(e), "detail": detail}), e.response.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/taxonomy/item-specifics", methods=["GET"])
def item_specifics():
    category_id = request.args.get("category_id", "").strip()
    if not category_id:
        return jsonify({"error": "Missing category_id parameter"}), 400

    marketplace = request.args.get("marketplace", "EBAY_GB")

    try:
        tree_id = get_category_tree_id(marketplace)
        token   = get_token()
        resp    = requests.get(
            f"{_ebay_base_url()}/commerce/taxonomy/v1/category_tree/{tree_id}/get_item_aspects_for_category",
            headers={"Authorization": f"Bearer {token}"},
            params={"category_id": category_id},
            timeout=15,
        )
        resp.raise_for_status()

        aspects = []
        for a in resp.json().get("aspects", []):
            constraint = a.get("aspectConstraint", {})
            required   = bool(constraint.get("aspectRequired"))
            usage      = constraint.get("aspectUsage", "OPTIONAL")
            level      = "REQUIRED" if required else ("RECOMMENDED" if usage == "RECOMMENDED" else "OPTIONAL")

            aspects.append({
                "name":                   a.get("localizedAspectName", ""),
                "requirementLevel":       level,
                "dataType":               constraint.get("aspectDataType", ""),
                "mode":                   constraint.get("aspectMode", ""),
                "supportsMultipleValues": constraint.get("itemToAspectCardinality") == "MULTI",
                "enabledForVariations":   bool(constraint.get("aspectEnabledForVariations")),
                "allowedValues":          [v.get("localizedValue", "") for v in a.get("aspectValues", [])],
                "searchCount":            a.get("relevanceIndicator", {}).get("searchCount", 0),
            })

        aspects.sort(key=lambda a: REQUIREMENT_ORDER.get(a["requirementLevel"], 3))

        summary = {
            "total":       len(aspects),
            "required":    sum(1 for a in aspects if a["requirementLevel"] == "REQUIRED"),
            "recommended": sum(1 for a in aspects if a["requirementLevel"] == "RECOMMENDED"),
            "optional":    sum(1 for a in aspects if a["requirementLevel"] == "OPTIONAL"),
        }

        return jsonify({"aspects": aspects, "summary": summary})
    except requests.HTTPError as e:
        try:
            detail = e.response.json()
        except Exception:
            detail = {"raw": e.response.text}
        return jsonify({"error": str(e), "detail": detail}), e.response.status_code
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
# Item Specifics Library — bulk taxonomy fetch + persistence
#
# Category data is written both as individual JSON files under
# data/item-specifics/<group>/<file>.json (for organised browsing on disk) and
# as one consolidated data/item_specifics_data.json. The in-memory cache is
# always the source of truth for the API — if the host's filesystem is
# ephemeral (e.g. wiped on redeploy) or read-only, disk writes fail silently
# and the feature keeps working from memory until the next restart, at which
# point load_item_specifics_cache() reloads whatever did make it to disk.
# ═══════════════════════════════════════════════════════════════════════════════

DATA_DIR                     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
ITEM_SPECIFICS_DIR           = os.path.join(DATA_DIR, "item-specifics")
ITEM_SPECIFICS_FALLBACK_FILE = os.path.join(DATA_DIR, "item_specifics_data.json")

_item_specifics_cache = {}   # "group/filename" -> record dict
_item_specifics_lock  = threading.Lock()

CATEGORY_MAP = [
    {"searchTerm": "Brake disc rotors",      "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-disc-rotors"},
    {"searchTerm": "Brake pads",             "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-pads"},
    {"searchTerm": "Brake shoes",            "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-shoes"},
    {"searchTerm": "Brake hoses",            "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-hoses"},
    {"searchTerm": "Brake cables",           "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-cables"},
    {"searchTerm": "Brake component kits",   "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-component-kits"},
    {"searchTerm": "Brake drums",            "group": "brakes",              "groupLabel": "Brakes",               "filename": "brake-drums"},

    {"searchTerm": "Control arms wishbones", "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "control-arms-wishbones"},
    {"searchTerm": "Ball joints",            "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "ball-joints"},
    {"searchTerm": "Tie rod ends",           "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "tie-rods"},
    {"searchTerm": "Stabiliser links",       "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "stabiliser-links"},
    {"searchTerm": "Bushes mountings",       "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "bushes-mountings"},
    {"searchTerm": "Shock absorbers",        "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "shocks-struts"},
    {"searchTerm": "Coil springs",           "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "springs"},
    {"searchTerm": "Wheel bearing kits",     "group": "steering-suspension", "groupLabel": "Steering & Suspension", "filename": "wheel-bearings"},

    {"searchTerm": "Engine gaskets seals",   "group": "engine",              "groupLabel": "Engine",               "filename": "engine-gaskets-seals"},
    {"searchTerm": "Camshaft parts",         "group": "engine",              "groupLabel": "Engine",               "filename": "camshaft-valve-parts"},
    {"searchTerm": "Timing belt kits",       "group": "engine",              "groupLabel": "Engine",               "filename": "timing-components"},
    {"searchTerm": "Accessory belts",        "group": "engine",              "groupLabel": "Engine",               "filename": "accessory-belts"},
    {"searchTerm": "Pulleys tensioners",     "group": "engine",              "groupLabel": "Engine",               "filename": "pulleys-tensioners"},
    {"searchTerm": "Thermostats",            "group": "engine",              "groupLabel": "Engine",               "filename": "thermostats"},
    {"searchTerm": "Water pumps",            "group": "engine",              "groupLabel": "Engine",               "filename": "water-pumps"},
    {"searchTerm": "Sensors",                "group": "engine",              "groupLabel": "Engine",               "filename": "sensors"},

    {"searchTerm": "Fuel injectors",         "group": "fuel-ignition",       "groupLabel": "Fuel & Ignition",      "filename": "fuel-injectors"},
    {"searchTerm": "Fuel pumps",             "group": "fuel-ignition",       "groupLabel": "Fuel & Ignition",      "filename": "fuel-pumps"},
    {"searchTerm": "Fuel filters",           "group": "fuel-ignition",       "groupLabel": "Fuel & Ignition",      "filename": "fuel-filters"},
    {"searchTerm": "Spark plugs",            "group": "fuel-ignition",       "groupLabel": "Fuel & Ignition",      "filename": "spark-plugs"},
    {"searchTerm": "Glow plugs",             "group": "fuel-ignition",       "groupLabel": "Fuel & Ignition",      "filename": "glow-plugs"},
    {"searchTerm": "Ignition coils leads",   "group": "fuel-ignition",       "groupLabel": "Fuel & Ignition",      "filename": "ignition-coils-leads"},

    {"searchTerm": "Air filters",            "group": "filters",             "groupLabel": "Filters",              "filename": "air-filters"},
    {"searchTerm": "Oil filters",            "group": "filters",             "groupLabel": "Filters",              "filename": "oil-filters"},
    {"searchTerm": "Cabin filters",          "group": "filters",             "groupLabel": "Filters",              "filename": "cabin-filters"},

    {"searchTerm": "Clutch kits",            "group": "transmission",        "groupLabel": "Transmission",         "filename": "clutch-parts"},
    {"searchTerm": "CV joints",              "group": "transmission",        "groupLabel": "Transmission",         "filename": "cv-joints"},
    {"searchTerm": "Propshafts",             "group": "transmission",        "groupLabel": "Transmission",         "filename": "propshafts"},

    {"searchTerm": "Alternators",            "group": "electrical",          "groupLabel": "Electrical",           "filename": "alternators"},
    {"searchTerm": "Starter motors",         "group": "electrical",          "groupLabel": "Electrical",           "filename": "starter-motors"},

    {"searchTerm": "Radiators",              "group": "cooling",             "groupLabel": "Cooling",              "filename": "radiators-cooling"},

    {"searchTerm": "EGR valves",             "group": "exhaust",             "groupLabel": "Exhaust",              "filename": "egr-emissions"},

    {"searchTerm": "Wiper blades",           "group": "exterior",            "groupLabel": "Exterior",             "filename": "wiper-blades"},
    {"searchTerm": "Headlights bulbs",       "group": "exterior",            "groupLabel": "Exterior",             "filename": "headlights-bulbs"},

    {"searchTerm": "Engine oil",             "group": "oils-fluids",         "groupLabel": "Oils & Fluids",        "filename": "engine-oil"},

    {"searchTerm": "Car care cleaning",      "group": "car-care",            "groupLabel": "Car Care",             "filename": "car-care-cleaning"},

    {"searchTerm": "AC compressors",         "group": "ac",                  "groupLabel": "Air Conditioning",     "filename": "ac-compressors"},
]

CATEGORY_MAP_BY_TERM = {c["searchTerm"].lower(): c for c in CATEGORY_MAP}

GROUP_ORDER = []
for _c in CATEGORY_MAP:
    if _c["group"] not in GROUP_ORDER:
        GROUP_ORDER.append(_c["group"])


def _slugify(text):
    s = re.sub(r"[^a-z0-9]+", "-", text.strip().lower()).strip("-")
    return s or "category"


def _cache_key(group, filename):
    return f"{group}/{filename}"


def load_item_specifics_cache():
    loaded = 0
    if os.path.isdir(ITEM_SPECIFICS_DIR):
        for group in os.listdir(ITEM_SPECIFICS_DIR):
            group_dir = os.path.join(ITEM_SPECIFICS_DIR, group)
            if not os.path.isdir(group_dir):
                continue
            for fname in os.listdir(group_dir):
                if not fname.endswith(".json"):
                    continue
                try:
                    with open(os.path.join(group_dir, fname), "r", encoding="utf-8") as f:
                        record = json.load(f)
                    _item_specifics_cache[_cache_key(group, fname[:-5])] = record
                    loaded += 1
                except (OSError, ValueError):
                    continue

    if loaded == 0 and os.path.exists(ITEM_SPECIFICS_FALLBACK_FILE):
        try:
            with open(ITEM_SPECIFICS_FALLBACK_FILE, "r", encoding="utf-8") as f:
                _item_specifics_cache.update(json.load(f))
        except (OSError, ValueError):
            pass


def persist_item_specifics_record(group, filename, record):
    _item_specifics_cache[_cache_key(group, filename)] = record
    try:
        group_dir = os.path.join(ITEM_SPECIFICS_DIR, group)
        os.makedirs(group_dir, exist_ok=True)
        with open(os.path.join(group_dir, f"{filename}.json"), "w", encoding="utf-8") as f:
            json.dump(record, f, indent=2)
    except OSError:
        pass  # ephemeral/read-only filesystem — in-memory cache still serves the API


def persist_item_specifics_fallback_file():
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(ITEM_SPECIFICS_FALLBACK_FILE, "w", encoding="utf-8") as f:
            json.dump(_item_specifics_cache, f, indent=2)
    except OSError:
        pass


load_item_specifics_cache()

_bulk_fetch_status = {
    "running":     False,
    "total":       0,
    "completed":   0,
    "current":     None,
    "results":     [],
    "started_at":  None,
    "finished_at": None,
}


def _fetch_one_category(term, marketplace):
    mapping = CATEGORY_MAP_BY_TERM.get(term.lower())
    if mapping:
        group, group_label, filename = mapping["group"], mapping["groupLabel"], mapping["filename"]
    else:
        group, group_label, filename = "custom", "Custom", _slugify(term)

    tree_id = get_category_tree_id(marketplace)
    token   = get_token()

    cat_resp = requests.get(
        f"{_ebay_base_url()}/commerce/taxonomy/v1/category_tree/{tree_id}/get_category_suggestions",
        headers={"Authorization": f"Bearer {token}"},
        params={"q": term},
        timeout=15,
    )
    cat_resp.raise_for_status()
    suggestions = cat_resp.json().get("categorySuggestions", [])
    if not suggestions:
        raise ValueError(f"No matching eBay category found for '{term}'")

    best          = suggestions[0]
    category      = best.get("category", {})
    category_id   = category.get("categoryId")
    category_name = category.get("categoryName", "")
    path          = _build_breadcrumb(category_name, best.get("categoryTreeNodeAncestors", []))

    aspects_resp = requests.get(
        f"{_ebay_base_url()}/commerce/taxonomy/v1/category_tree/{tree_id}/get_item_aspects_for_category",
        headers={"Authorization": f"Bearer {token}"},
        params={"category_id": category_id},
        timeout=15,
    )
    aspects_resp.raise_for_status()

    aspects = []
    for a in aspects_resp.json().get("aspects", []):
        constraint = a.get("aspectConstraint", {})
        required   = bool(constraint.get("aspectRequired"))
        usage      = constraint.get("aspectUsage", "OPTIONAL")
        level      = "REQUIRED" if required else ("RECOMMENDED" if usage == "RECOMMENDED" else "OPTIONAL")

        aspects.append({
            "name":                 a.get("localizedAspectName", ""),
            "level":                level,
            "dataType":             constraint.get("aspectDataType", ""),
            "mode":                 constraint.get("aspectMode", ""),
            "multiValue":           constraint.get("itemToAspectCardinality") == "MULTI",
            "enabledForVariations": bool(constraint.get("aspectEnabledForVariations")),
            "values":               [v.get("localizedValue", "") for v in a.get("aspectValues", [])],
            "searchCount":          a.get("relevanceIndicator", {}).get("searchCount", 0),
        })
    aspects.sort(key=lambda a: REQUIREMENT_ORDER.get(a["level"], 3))

    counts = {
        "total":       len(aspects),
        "required":    sum(1 for a in aspects if a["level"] == "REQUIRED"),
        "recommended": sum(1 for a in aspects if a["level"] == "RECOMMENDED"),
        "optional":    sum(1 for a in aspects if a["level"] == "OPTIONAL"),
    }

    record = {
        "searchTerm":  term,
        "group":       group,
        "groupLabel":  group_label,
        "filename":    filename,
        "ebayCategory": {
            "categoryId":   category_id,
            "categoryName": category_name,
            "path":         path,
        },
        "marketplace": marketplace,
        "fetchedAt":   datetime.now(timezone.utc).isoformat(),
        "counts":      counts,
        "aspects":     aspects,
    }
    return group, filename, record


def _run_bulk_fetch(terms, marketplace):
    results = []
    for i, term in enumerate(terms):
        _bulk_fetch_status["current"] = term
        try:
            group, filename, record = _fetch_one_category(term, marketplace)
            with _item_specifics_lock:
                persist_item_specifics_record(group, filename, record)
            results.append({
                "searchTerm":   term,
                "status":       "ok",
                "group":        group,
                "filename":     filename,
                "categoryName": record["ebayCategory"]["categoryName"],
                "counts":       record["counts"],
            })
        except requests.HTTPError as e:
            try:
                detail = e.response.json()
            except Exception:
                detail = {"raw": e.response.text}
            results.append({"searchTerm": term, "status": "error", "error": str(e), "detail": detail})
        except Exception as e:
            results.append({"searchTerm": term, "status": "error", "error": str(e)})

        _bulk_fetch_status["completed"] = i + 1
        _bulk_fetch_status["results"]   = list(results)

        if i < len(terms) - 1:
            time.sleep(1)

    with _item_specifics_lock:
        persist_item_specifics_fallback_file()

    _bulk_fetch_status.update({
        "running":     False,
        "current":     None,
        "finished_at": datetime.now(timezone.utc).isoformat(),
    })


@app.route("/api/taxonomy/bulk-fetch", methods=["POST"])
def bulk_fetch_categories():
    if _bulk_fetch_status["running"]:
        return jsonify({"error": "A bulk fetch is already in progress"}), 409

    data      = request.get_json(silent=True) or {}
    raw_terms = data.get("searchTerms")
    if raw_terms:
        if not isinstance(raw_terms, list):
            return jsonify({"error": "searchTerms must be a list"}), 400
        terms = [str(t).strip() for t in raw_terms if str(t).strip()]
    else:
        terms = [c["searchTerm"] for c in CATEGORY_MAP]

    if not terms:
        return jsonify({"error": "searchTerms must be a non-empty list"}), 400

    marketplace = data.get("marketplace", "EBAY_GB")

    _bulk_fetch_status.update({
        "running":     True,
        "total":       len(terms),
        "completed":   0,
        "current":     None,
        "results":     [],
        "started_at":  datetime.now(timezone.utc).isoformat(),
        "finished_at": None,
    })

    threading.Thread(target=_run_bulk_fetch, args=(terms, marketplace), daemon=True).start()

    return jsonify({"started": True, "total": len(terms)}), 202


@app.route("/api/taxonomy/bulk-fetch/status", methods=["GET"])
def bulk_fetch_status():
    return jsonify(_bulk_fetch_status)


@app.route("/api/taxonomy/saved-categories", methods=["GET"])
def saved_categories():
    groups = {}
    for record in _item_specifics_cache.values():
        group       = record.get("group", "custom")
        group_label = record.get("groupLabel") or group.replace("-", " ").title()
        groups.setdefault(group, {"key": group, "label": group_label, "categories": []})
        groups[group]["categories"].append(record)

    ordered_keys  = [g for g in GROUP_ORDER if g in groups]
    ordered_keys += sorted(g for g in groups if g not in GROUP_ORDER)

    result = []
    for g in ordered_keys:
        entry = groups[g]
        entry["categories"].sort(key=lambda c: c.get("searchTerm", ""))
        result.append(entry)

    return jsonify({"groups": result})


# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
